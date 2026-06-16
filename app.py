import os
import re
import unicodedata


import openpyxl
from flask import Flask, jsonify, render_template, request, send_file

app = Flask(__name__)

# ---------------------------------------------------------------------------
# Configuracion  -  ajusta estas variables
# ---------------------------------------------------------------------------
HOME = os.path.expanduser("~")
if os.path.exists(os.path.join(HOME, "certificados-web")):
    BASE = os.path.join(HOME, "certificados-web")
else:
    BASE = os.path.join(HOME, "Desktop", "Colegios Unidos", "certificados-web")

EXCEL_PATH = os.path.join(BASE, "data", "contactos.xlsx")
CERT_BASE = os.path.join(BASE, "certificados")


def normalizar(texto):
    s = unicodedata.normalize("NFKD", texto).encode("ASCII", "ignore").decode()
    s = re.sub(r"[^\w\s]", "", s)
    return " ".join(s.lower().split())


def cargar_certificados():
    mapa = {}
    if not os.path.isdir(CERT_BASE):
        return mapa
    for f in os.listdir(CERT_BASE):
        if not f.endswith(".jpg"):
            continue
        nombre = f[:-4]
        mapa[normalizar(nombre)] = {
            "path": os.path.join(CERT_BASE, f),
        }
    return mapa


def cargar_participantes(mapa_certs):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    personas = []
    for row in range(2, ws.max_row + 1):
        nombre = (ws.cell(row, 1).value or "").strip()
        apellido = (ws.cell(row, 2).value or "").strip()
        email = (ws.cell(row, 3).value or "").strip().lower()
        if not email:
            continue
        nombre_completo = f"{nombre} {apellido}".strip()
        norm = normalizar(nombre_completo)
        cert = mapa_certs.get(norm)
        personas.append({
            "nombre": nombre_completo,
            "email": email,
            "certificado": cert["path"] if cert else None,
        })
    return personas


CERT_MAP = cargar_certificados()
PERSONAS = cargar_participantes(CERT_MAP)
EMAIL_INDEX = {p["email"]: p for p in PERSONAS if p.get("certificado")}


def obtener_persona(email):
    return EMAIL_INDEX.get(email.strip().lower())


# ---------------------------------------------------------------------------
# LinkedIn share URL
# ---------------------------------------------------------------------------
def linkedin_share_url(persona):
    texto = (
        f"Me capacity\u00e9 en el 2\u00b0 Encuentro de Colegios Unidos "
        f"sobre IA y bienestar en el mapa escolar. "
        f"\u00a1Gracias por esta gran experiencia!"
    )
    from urllib.parse import quote
    return f"https://www.linkedin.com/sharing/share-offsite/?url=https://colegiosunidos.org&summary={quote(texto)}"


# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/verificar", methods=["POST"])
def verificar():
    data = request.get_json()
    email = (data.get("email") or "").strip().lower()
    persona = obtener_persona(email)
    if not persona:
        return jsonify({
            "existe": False,
            "error": "Oh lo siento, tu certificado no fue encontrado. Te recomendamos ponerte en contacto con los organizadores."
        }), 404
    return jsonify({
        "existe": True,
        "nombre": persona["nombre"],
    })


@app.route("/api/certificado")
def certificado():
    email = request.args.get("email", "").strip().lower()
    persona = obtener_persona(email)
    if not persona or not persona.get("certificado"):
        return jsonify({"error": "No encontrado"}), 404
    if request.args.get("download"):
        return send_file(persona["certificado"], mimetype="image/jpeg", as_attachment=True, download_name=f"certificado_{persona['nombre'].replace(' ','_')}.jpg")
    return send_file(persona["certificado"], mimetype="image/jpeg")



@app.route("/api/share-url")
def share_url():
    email = request.args.get("email", "").strip().lower()
    persona = obtener_persona(email)
    if not persona:
        return jsonify({"error": "No encontrado"}), 404
    return jsonify({"url": linkedin_share_url(persona)})


if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    print(f"Participantes cargados: {len(PERSONAS)}")
    print(f"Con certificado asignado: {sum(1 for p in PERSONAS if p.get('certificado'))}")
    app.run(host="0.0.0.0", port=port)
